from __future__ import annotations

import calendar
import json
import sqlite3
import uuid
from datetime import date, datetime, timezone
from io import BytesIO
from typing import Any

from flask import Blueprint, Flask, current_app, jsonify, redirect, render_template, request, send_file, session
from openpyxl import Workbook
from openpyxl.styles import Font

from ai_service import AIServiceError
from db import get_account, get_db
from db_backend import column_names as backend_column_names
from finance_ai import FinanceIntent, fold_text, parse_finance_message

bp = Blueprint("finance", __name__)

EXPENSE_CATEGORIES = {
    # Existing personal categories stay valid for old data.
    "food": "Ăn uống",
    "housing": "Nhà ở",
    "transport": "Đi lại",
    "shopping": "Mua sắm",
    "study": "Học tập",
    "health": "Sức khỏe",
    "entertainment": "Giải trí",
    "bills": "Hóa đơn",
    "family": "Gia đình",
    # Company-oriented categories.
    "client_entertainment": "Tiếp khách",
    "travel": "Công tác",
    "office": "Văn phòng phẩm",
    "equipment": "Thiết bị / công cụ",
    "software": "Phần mềm / dịch vụ số",
    "telecom": "Internet / viễn thông",
    "rent": "Thuê văn phòng / mặt bằng",
    "utilities": "Điện nước / tiện ích",
    "marketing": "Marketing / quảng cáo",
    "shipping": "Vận chuyển / logistics",
    "maintenance": "Bảo trì / sửa chữa",
    "payroll": "Lương nhân viên",
    "tax": "Thuế / lệ phí",
    "training": "Đào tạo",
    "other": "Khác",
}
INCOME_CATEGORIES = {
    "salary": "Lương",
    "freelance": "Freelance",
    "bonus": "Thưởng",
    "family_support": "Gia đình hỗ trợ",
    "sale": "Doanh thu / khách thanh toán",
    "other_income": "Thu nhập khác",
}
ALL_CATEGORIES = {**EXPENSE_CATEGORIES, **INCOME_CATEGORIES}

AI_AUTOSAVE_CONFIDENCE = 0.78


def _now() -> str:
    return datetime.now(timezone.utc).isoformat()


def _user_id() -> str | None:
    value = str(session.get("account_id", "")).strip()
    return value or None


def _error(message: str, status: int = 400, code: str = "bad_request"):
    return jsonify({"error": message, "code": code}), status


def _parse_month(value: str | None) -> str:
    raw = str(value or "").strip()
    if not raw:
        return date.today().strftime("%Y-%m")
    try:
        datetime.strptime(raw, "%Y-%m")
    except ValueError as exc:
        raise ValueError("Tháng không hợp lệ. Dùng định dạng YYYY-MM.") from exc
    return raw


def _month_bounds(month: str) -> tuple[str, str]:
    year, month_num = [int(x) for x in month.split("-")]
    last_day = calendar.monthrange(year, month_num)[1]
    return f"{month}-01", f"{month}-{last_day:02d}"


def _amount(value: Any, *, allow_zero: bool = False) -> int:
    try:
        number = int(round(float(value)))
    except (TypeError, ValueError) as exc:
        raise ValueError("Số tiền không hợp lệ.") from exc
    if allow_zero:
        if number < 0:
            raise ValueError("Số tiền không được âm.")
    elif number <= 0:
        raise ValueError("Số tiền phải lớn hơn 0.")
    if number > 10_000_000_000_000:
        raise ValueError("Số tiền vượt quá giới hạn cho phép.")
    return number


def _column_names(db, table: str) -> set[str]:
    return backend_column_names(db, table)


def init_finance_db() -> None:
    db = get_db()
    db.executescript(
        """
        CREATE TABLE IF NOT EXISTS finance_transactions (
            id TEXT PRIMARY KEY,
            user_id TEXT NOT NULL,
            kind TEXT NOT NULL CHECK(kind IN ('income','expense')),
            amount INTEGER NOT NULL CHECK(amount > 0),
            category TEXT NOT NULL,
            note TEXT NOT NULL DEFAULT '',
            occurred_on TEXT NOT NULL,
            created_at TEXT NOT NULL,
            FOREIGN KEY(user_id) REFERENCES users(id) ON DELETE CASCADE
        );

        CREATE INDEX IF NOT EXISTS idx_finance_transactions_user_date
        ON finance_transactions(user_id, occurred_on DESC, created_at DESC);

        CREATE TABLE IF NOT EXISTS finance_monthly_plans (
            user_id TEXT NOT NULL,
            month TEXT NOT NULL,
            monthly_income_target INTEGER NOT NULL DEFAULT 0,
            budget_limit INTEGER NOT NULL DEFAULT 0,
            saving_target INTEGER NOT NULL DEFAULT 0,
            updated_at TEXT NOT NULL,
            PRIMARY KEY(user_id, month),
            FOREIGN KEY(user_id) REFERENCES users(id) ON DELETE CASCADE
        );

        CREATE TABLE IF NOT EXISTS finance_chat_messages (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            user_id TEXT NOT NULL,
            role TEXT NOT NULL CHECK(role IN ('user','assistant')),
            content TEXT NOT NULL,
            meta_json TEXT NOT NULL DEFAULT '{}',
            created_at TEXT NOT NULL,
            FOREIGN KEY(user_id) REFERENCES users(id) ON DELETE CASCADE
        );

        CREATE INDEX IF NOT EXISTS idx_finance_chat_user_id
        ON finance_chat_messages(user_id, id DESC);

        CREATE TABLE IF NOT EXISTS finance_assistant_state (
            user_id TEXT PRIMARY KEY,
            pending_json TEXT NOT NULL DEFAULT '',
            last_transaction_id TEXT NOT NULL DEFAULT '',
            updated_at TEXT NOT NULL,
            FOREIGN KEY(user_id) REFERENCES users(id) ON DELETE CASCADE
        );
        """
    )

    columns = _column_names(db, "finance_transactions")
    migrations = {
        "currency": "TEXT NOT NULL DEFAULT 'VND'",
        "vendor": "TEXT NOT NULL DEFAULT ''",
        "department": "TEXT NOT NULL DEFAULT ''",
        "project": "TEXT NOT NULL DEFAULT ''",
        "client": "TEXT NOT NULL DEFAULT ''",
        "payment_method": "TEXT NOT NULL DEFAULT ''",
        "source": "TEXT NOT NULL DEFAULT 'manual'",
        "ai_confidence": "REAL NOT NULL DEFAULT 1.0",
        "review_status": "TEXT NOT NULL DEFAULT 'confirmed'",
    }
    for name, ddl in migrations.items():
        if name not in columns:
            db.execute(f"ALTER TABLE finance_transactions ADD COLUMN {name} {ddl}")
    db.commit()


def register_finance(app: Flask) -> None:
    with app.app_context():
        init_finance_db()
    app.register_blueprint(bp)


def _serialize_transaction(row: sqlite3.Row | dict[str, Any]) -> dict[str, Any]:
    item = dict(row)
    item["amount"] = int(item.get("amount") or 0)
    item["ai_confidence"] = float(item.get("ai_confidence") or 0)
    return item


def _transaction_select() -> str:
    return (
        "id, kind, amount, category, note, occurred_on, created_at, currency, vendor, "
        "department, project, client, payment_method, source, ai_confidence, review_status"
    )


def _plan(user_id: str, month: str) -> dict[str, Any]:
    row = get_db().execute(
        """
        SELECT monthly_income_target, budget_limit, saving_target, updated_at
        FROM finance_monthly_plans
        WHERE user_id = ? AND month = ?
        """,
        (user_id, month),
    ).fetchone()
    if not row:
        return {
            "month": month,
            "monthly_income_target": 0,
            "budget_limit": 0,
            "saving_target": 0,
            "updated_at": "",
        }
    item = dict(row)
    item["month"] = month
    for key in ("monthly_income_target", "budget_limit", "saving_target"):
        item[key] = int(item.get(key) or 0)
    return item


def _overview(user_id: str, plan_month: str | None = None, recent_limit: int = 20) -> dict[str, Any]:
    """Return all-time finance totals/categories plus the most recent transactions.

    Month is no longer a dashboard filter. ``plan_month`` is used only for the
    optional monthly budget plan so old planning data keeps working.
    """
    plan_month = _parse_month(plan_month or date.today().strftime("%Y-%m"))
    db = get_db()

    totals_row = db.execute(
        """
        SELECT
            COALESCE(SUM(CASE WHEN kind = 'income' THEN amount ELSE 0 END), 0) AS income,
            COALESCE(SUM(CASE WHEN kind = 'expense' THEN amount ELSE 0 END), 0) AS expense
        FROM finance_transactions
        WHERE user_id = ?
        """,
        (user_id,),
    ).fetchone()
    income = int(totals_row["income"] or 0) if totals_row else 0
    expense = int(totals_row["expense"] or 0) if totals_row else 0
    balance = income - expense

    category_rows = db.execute(
        """
        SELECT category, COALESCE(SUM(amount), 0) AS total
        FROM finance_transactions
        WHERE user_id = ? AND kind = 'expense'
        GROUP BY category
        ORDER BY total DESC
        """,
        (user_id,),
    ).fetchall()
    categories = [
        {
            "key": str(row["category"]),
            "label": EXPENSE_CATEGORIES.get(str(row["category"]), str(row["category"])),
            "amount": int(row["total"] or 0),
            "share": round((int(row["total"] or 0) / expense) * 100, 1) if expense else 0,
        }
        for row in category_rows
    ]

    recent_rows = db.execute(
        f"""
        SELECT {_transaction_select()}
        FROM finance_transactions
        WHERE user_id = ?
        ORDER BY created_at DESC, occurred_on DESC
        LIMIT ?
        """,
        (user_id, max(1, min(int(recent_limit), 100))),
    ).fetchall()
    transactions = [_serialize_transaction(row) for row in recent_rows]

    plan = _plan(user_id, plan_month)
    start, end = _month_bounds(plan_month)
    month_row = db.execute(
        """
        SELECT
            COALESCE(SUM(CASE WHEN kind = 'income' THEN amount ELSE 0 END), 0) AS income,
            COALESCE(SUM(CASE WHEN kind = 'expense' THEN amount ELSE 0 END), 0) AS expense
        FROM finance_transactions
        WHERE user_id = ? AND occurred_on BETWEEN ? AND ?
        """,
        (user_id, start, end),
    ).fetchone()
    month_income = int(month_row["income"] or 0) if month_row else 0
    month_expense = int(month_row["expense"] or 0) if month_row else 0
    month_balance = month_income - month_expense

    budget_limit = int(plan["budget_limit"] or 0)
    saving_target = int(plan["saving_target"] or 0)
    income_target = int(plan["monthly_income_target"] or 0)
    budget_used_percent = round((month_expense / budget_limit) * 100, 1) if budget_limit else 0
    saving_progress_percent = round((max(month_balance, 0) / saving_target) * 100, 1) if saving_target else 0
    income_progress_percent = round((month_income / income_target) * 100, 1) if income_target else 0

    status = "balanced"
    if budget_limit and month_expense > budget_limit:
        status = "over_budget"
    elif saving_target and month_balance < saving_target:
        status = "below_saving_target"
    elif saving_target and month_balance >= saving_target:
        status = "on_track"

    return {
        "scope": "all",
        "month": plan_month,
        "summary": {
            "income": income,
            "expense": expense,
            "balance": balance,
            "budget_remaining": max(budget_limit - month_expense, 0) if budget_limit else 0,
            "budget_over": max(month_expense - budget_limit, 0) if budget_limit else 0,
            "budget_used_percent": budget_used_percent,
            "saving_progress_percent": saving_progress_percent,
            "income_progress_percent": income_progress_percent,
            "status": status,
        },
        "plan": plan,
        "categories": categories,
        "transactions": transactions,
        "expense_categories": EXPENSE_CATEGORIES,
        "income_categories": INCOME_CATEGORIES,
    }


def _save_chat(user_id: str, role: str, content: str, meta: dict[str, Any] | None = None) -> dict[str, Any]:
    db = get_db()
    cursor = db.execute(
        "INSERT INTO finance_chat_messages(user_id, role, content, meta_json, created_at) VALUES (?, ?, ?, ?, ?) RETURNING id",
        (user_id, role, str(content)[:4000], json.dumps(meta or {}, ensure_ascii=False), _now()),
    )
    inserted = cursor.fetchone()
    db.commit()
    if not inserted:
        raise RuntimeError("Không lấy được id finance chat vừa tạo.")
    return {
        "id": int(inserted["id"]),
        "role": role,
        "content": str(content),
        "meta": meta or {},
    }


def _chat_history(user_id: str, limit: int = 60) -> list[dict[str, Any]]:
    rows = get_db().execute(
        "SELECT id, role, content, meta_json, created_at FROM finance_chat_messages WHERE user_id = ? ORDER BY id DESC LIMIT ?",
        (user_id, max(1, min(int(limit), 200))),
    ).fetchall()
    output: list[dict[str, Any]] = []
    for row in reversed(rows):
        try:
            meta = json.loads(str(row["meta_json"] or "{}"))
        except json.JSONDecodeError:
            meta = {}
        output.append({"id": int(row["id"]), "role": row["role"], "content": row["content"], "meta": meta, "created_at": row["created_at"]})
    return output


def _assistant_state(user_id: str) -> dict[str, Any]:
    row = get_db().execute(
        "SELECT pending_json, last_transaction_id FROM finance_assistant_state WHERE user_id = ?",
        (user_id,),
    ).fetchone()
    if not row:
        return {"pending": None, "last_transaction_id": ""}
    try:
        pending = json.loads(str(row["pending_json"] or "")) if str(row["pending_json"] or "").strip() else None
    except json.JSONDecodeError:
        pending = None
    return {"pending": pending, "last_transaction_id": str(row["last_transaction_id"] or "")}


def _save_state(user_id: str, *, pending: dict[str, Any] | None = None, last_transaction_id: str | None = None) -> None:
    current = _assistant_state(user_id)
    last_id = current["last_transaction_id"] if last_transaction_id is None else str(last_transaction_id or "")
    db = get_db()
    db.execute(
        """
        INSERT INTO finance_assistant_state(user_id, pending_json, last_transaction_id, updated_at)
        VALUES (?, ?, ?, ?)
        ON CONFLICT(user_id) DO UPDATE SET
            pending_json = excluded.pending_json,
            last_transaction_id = excluded.last_transaction_id,
            updated_at = excluded.updated_at
        """,
        (user_id, json.dumps(pending or {}, ensure_ascii=False) if pending else "", last_id, _now()),
    )
    db.commit()


def _clean_meta(value: Any, limit: int = 120) -> str:
    return " ".join(str(value or "").split()).strip()[:limit]


def _insert_transaction(user_id: str, draft: dict[str, Any], *, source: str) -> dict[str, Any]:
    kind = str(draft.get("kind") or "expense")
    categories = INCOME_CATEGORIES if kind == "income" else EXPENSE_CATEGORIES
    category = str(draft.get("category") or "")
    if category not in categories:
        raise ValueError("Danh mục không hợp lệ.")
    amount = _amount(draft.get("amount"))
    occurred_on = str(draft.get("occurred_on") or date.today().isoformat())
    datetime.strptime(occurred_on, "%Y-%m-%d")
    transaction_id = str(uuid.uuid4())
    now = _now()
    db = get_db()
    db.execute(
        """
        INSERT INTO finance_transactions(
            id, user_id, kind, amount, category, note, occurred_on, created_at,
            currency, vendor, department, project, client, payment_method,
            source, ai_confidence, review_status
        ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """,
        (
            transaction_id,
            user_id,
            kind,
            amount,
            category,
            _clean_meta(draft.get("note"), 500),
            occurred_on,
            now,
            _clean_meta(draft.get("currency") or "VND", 12),
            _clean_meta(draft.get("vendor")),
            _clean_meta(draft.get("department")),
            _clean_meta(draft.get("project")),
            _clean_meta(draft.get("client")),
            _clean_meta(draft.get("payment_method"), 40),
            source,
            max(0.0, min(1.0, float(draft.get("confidence") or 1.0))),
            _clean_meta(draft.get("review_status") or "confirmed", 40),
        ),
    )
    db.commit()
    row = db.execute(f"SELECT {_transaction_select()} FROM finance_transactions WHERE id = ? AND user_id = ?", (transaction_id, user_id)).fetchone()
    return _serialize_transaction(row)


def _update_last_transaction(user_id: str, transaction_id: str, fields: dict[str, Any]) -> dict[str, Any] | None:
    row = get_db().execute(
        f"SELECT {_transaction_select()} FROM finance_transactions WHERE id = ? AND user_id = ?",
        (transaction_id, user_id),
    ).fetchone()
    if not row:
        return None
    current = _serialize_transaction(row)
    allowed = {"kind", "amount", "category", "occurred_on", "note", "vendor", "department", "project", "client", "payment_method"}
    updates = {key: value for key, value in fields.items() if key in allowed and value not in {None, ""}}
    if not updates:
        return current
    if "amount" in updates:
        updates["amount"] = _amount(updates["amount"])
    if "occurred_on" in updates:
        datetime.strptime(str(updates["occurred_on"]), "%Y-%m-%d")
    if "category" in updates:
        kind = str(updates.get("kind") or current.get("kind") or "expense")
        categories = INCOME_CATEGORIES if kind == "income" else EXPENSE_CATEGORIES
        if str(updates["category"]) not in categories:
            updates.pop("category", None)
    if not updates:
        return current
    assignments = ", ".join(f"{key} = ?" for key in updates)
    values = list(updates.values()) + [transaction_id, user_id]
    db = get_db()
    db.execute(f"UPDATE finance_transactions SET {assignments} WHERE id = ? AND user_id = ?", values)
    db.commit()
    row = db.execute(f"SELECT {_transaction_select()} FROM finance_transactions WHERE id = ? AND user_id = ?", (transaction_id, user_id)).fetchone()
    return _serialize_transaction(row) if row else None


def _format_money(value: int) -> str:
    return f"{int(value):,}".replace(",", ".") + "đ"


def _transaction_summary(item: dict[str, Any]) -> str:
    label = ALL_CATEGORIES.get(str(item.get("category")), str(item.get("category") or "Khác"))
    kind_label = "Thu" if item.get("kind") == "income" else "Chi"
    extras = []
    if item.get("department"):
        extras.append(f"{item['department']}")
    if item.get("payment_method"):
        payment = {"bank_transfer": "chuyển khoản", "cash": "tiền mặt", "credit_card": "thẻ"}.get(item["payment_method"], item["payment_method"])
        extras.append(payment)
    suffix = f" · {' · '.join(extras)}" if extras else ""
    return f"Đã lưu: {kind_label} {_format_money(item['amount'])} · {label} · {item['occurred_on']}{suffix}."


def _build_export_data(transactions: list[dict[str, Any]]) -> dict[str, Any]:
    """Build export-ready finance data without exposing internal source metadata."""
    income = sum(int(item.get("amount") or 0) for item in transactions if item.get("kind") == "income")
    expense = sum(int(item.get("amount") or 0) for item in transactions if item.get("kind") == "expense")
    category_amounts: dict[str, int] = {}
    for item in transactions:
        if item.get("kind") != "expense":
            continue
        key = str(item.get("category") or "other")
        category_amounts[key] = category_amounts.get(key, 0) + int(item.get("amount") or 0)
    category_totals = [
        (key, EXPENSE_CATEGORIES.get(key, key), amount)
        for key, amount in sorted(category_amounts.items(), key=lambda pair: pair[1], reverse=True)
    ]
    headers = ["Ngày", "Loại", "Số tiền", "Danh mục", "Nội dung", "Nhà cung cấp", "Phòng ban", "Dự án", "Khách hàng", "Thanh toán"]
    detail_rows = [
        [
            item.get("occurred_on", ""),
            "Thu" if item.get("kind") == "income" else "Chi",
            int(item.get("amount") or 0),
            ALL_CATEGORIES.get(str(item.get("category") or ""), str(item.get("category") or "")),
            item.get("note", ""),
            item.get("vendor", ""),
            item.get("department", ""),
            item.get("project", ""),
            item.get("client", ""),
            item.get("payment_method", ""),
        ]
        for item in transactions
    ]
    return {
        "summary": {"income": income, "expense": expense, "balance": income - expense},
        "category_totals": category_totals,
        "detail_headers": headers,
        "detail_rows": detail_rows,
    }


def _merge_pending(pending: dict[str, Any], followup: FinanceIntent) -> dict[str, Any]:
    merged = dict(pending)
    for key in ("kind", "amount", "category", "occurred_on", "department", "payment_method", "vendor", "project", "client"):
        value = getattr(followup, key, None)
        if value not in {None, ""}:
            # Avoid a standalone follow-up defaulting an old explicit date back to today unless it contains date words.
            if key == "occurred_on" and value == date.today().isoformat():
                folded = fold_text(followup.note)
                if not any(marker in folded for marker in ("hom nay", "hom qua", "hom kia", "/", "-")):
                    continue
            merged[key] = value
    if followup.note:
        merged["note"] = str(merged.get("note") or "") + " | " + followup.note
    return merged


def _pending_question(draft: dict[str, Any]) -> str:
    if not draft.get("amount"):
        return "Khoản này bao nhiêu tiền?"
    if not draft.get("occurred_on"):
        return "Khoản này chính xác là ngày nào?"
    if not draft.get("category"):
        return "Tao chưa chắc nên xếp khoản này vào danh mục nào. Mày nói ngắn kiểu “văn phòng phẩm”, “tiếp khách”, “đi lại”… là được."
    return "Tao còn thiếu một dữ kiện để lưu khoản này."


def _query_total(user_id: str, intent: FinanceIntent) -> tuple[int, int]:
    sql = "SELECT COALESCE(SUM(amount),0) AS total, COUNT(*) AS n FROM finance_transactions WHERE user_id = ?"
    params: list[Any] = [user_id]
    if intent.month:
        start, end = _month_bounds(intent.month)
        sql += " AND occurred_on BETWEEN ? AND ?"
        params.extend([start, end])
    if intent.kind in {"income", "expense"}:
        sql += " AND kind = ?"
        params.append(intent.kind)
    if intent.category:
        sql += " AND category = ?"
        params.append(intent.category)
    row = get_db().execute(sql, params).fetchone()
    return int(row["total"] or 0), int(row["n"] or 0)


def _finance_ai_classifier(message: str, intent: FinanceIntent) -> dict[str, Any] | None:
    ai = current_app.extensions.get("ai_service") if hasattr(current_app, "extensions") else None
    if not ai or not getattr(ai, "is_configured", False) or not hasattr(ai, "classify_finance_transaction"):
        return None
    categories = INCOME_CATEGORIES if intent.kind == "income" else EXPENSE_CATEGORIES
    try:
        return ai.classify_finance_transaction(message, allowed_categories=categories)
    except AIServiceError:
        current_app.logger.warning("Finance Luna classification failed", exc_info=True)
        return None


def _looks_like_last_edit(message: str) -> bool:
    folded = fold_text(message)
    return any(
        folded.startswith(marker)
        for marker in ("a ", "ah ", "hom qua", "hom kia", "doi sang", "doi thanh", "sua thanh", "chuyen sang", "phong ")
    )


@bp.get("/finance")
def finance_page():
    user_id = _user_id()
    if not user_id or not get_account(user_id):
        return redirect("/")
    account = get_account(user_id) or {}
    return render_template("finance/index.html", display_name=account.get("display_name", "Bạn"))


@bp.get("/api/finance/overview")
def finance_overview():
    user_id = _user_id()
    if not user_id:
        return _error("Bạn cần đăng nhập trước.", 401, "auth_required")
    try:
        plan_month = _parse_month(request.args.get("plan_month")) if request.args.get("plan_month") else date.today().strftime("%Y-%m")
    except ValueError as exc:
        return _error(str(exc))
    return jsonify(_overview(user_id, plan_month=plan_month))


@bp.get("/api/finance/chat")
def finance_chat_history():
    user_id = _user_id()
    if not user_id:
        return _error("Bạn cần đăng nhập trước.", 401, "auth_required")
    state = _assistant_state(user_id)
    pending = state.get("pending")
    return jsonify({"state": state, "pending_prompt": _pending_question(pending) if pending else ""})


@bp.post("/api/finance/assistant")
def finance_assistant():
    user_id = _user_id()
    if not user_id:
        return _error("Bạn cần đăng nhập trước.", 401, "auth_required")
    payload = request.get_json(silent=True) or {}
    message = " ".join(str(payload.get("message") or "").split()).strip()
    if not message:
        return _error("Hãy nhập một khoản chi hoặc một lệnh.")
    if len(message) > 1000:
        return _error("Tin nhắn quản lý chi tiêu tối đa 1000 ký tự.")
    _save_chat(user_id, "user", message)
    state = _assistant_state(user_id)
    intent = parse_finance_message(message)

    # Pending draft: the user's next message fills only missing fields, then save as soon as it is complete.
    if state.get("pending"):
        pending = _merge_pending(dict(state["pending"]), intent)
        if pending.get("amount") and pending.get("occurred_on") and pending.get("category"):
            pending["confidence"] = float(pending.get("confidence") or 1.0)
            pending["review_status"] = "confirmed"
            item = _insert_transaction(user_id, pending, source="finance_chat")
            _save_state(user_id, pending=None, last_transaction_id=item["id"])
            reply = _transaction_summary(item)
            assistant = _save_chat(user_id, "assistant", reply, {"action": "created", "transaction_id": item["id"], "used_ai": False})
            return jsonify({"reply": reply, "assistant_message": assistant, "action": "created", "transaction": item, "overview": _overview(user_id), "used_ai": False})
        _save_state(user_id, pending=pending)
        reply = _pending_question(pending)
        assistant = _save_chat(user_id, "assistant", reply, {"action": "clarify", "draft": pending})
        return jsonify({"reply": reply, "assistant_message": assistant, "action": "clarify", "draft": pending, "used_ai": False})

    if intent.action == "delete_last":
        last_id = str(state.get("last_transaction_id") or "")
        if not last_id:
            reply = "Chưa có khoản gần nhất nào để xóa."
            assistant = _save_chat(user_id, "assistant", reply, {"action": "noop"})
            return jsonify({"reply": reply, "assistant_message": assistant, "action": "noop"})
        row = get_db().execute("SELECT occurred_on FROM finance_transactions WHERE id = ? AND user_id = ?", (last_id, user_id)).fetchone()
        if not row:
            _save_state(user_id, last_transaction_id="")
            reply = "Khoản gần nhất không còn tồn tại."
            assistant = _save_chat(user_id, "assistant", reply, {"action": "noop"})
            return jsonify({"reply": reply, "assistant_message": assistant, "action": "noop"})
        month = str(row["occurred_on"])[:7]
        get_db().execute("DELETE FROM finance_transactions WHERE id = ? AND user_id = ?", (last_id, user_id))
        get_db().commit()
        _save_state(user_id, last_transaction_id="")
        reply = "Đã xóa khoản vừa rồi."
        assistant = _save_chat(user_id, "assistant", reply, {"action": "deleted", "transaction_id": last_id})
        return jsonify({"reply": reply, "assistant_message": assistant, "action": "deleted", "overview": _overview(user_id), "used_ai": False})

    if intent.action == "query":
        if intent.kind not in {"income", "expense"} and not intent.category:
            sql = "SELECT kind, COALESCE(SUM(amount),0) AS total, COUNT(*) AS n FROM finance_transactions WHERE user_id = ?"
            params: list[Any] = [user_id]
            scope_label = "Tổng cộng"
            if intent.month:
                start, end = _month_bounds(intent.month)
                sql += " AND occurred_on BETWEEN ? AND ?"
                params.extend([start, end])
                scope_label = f"Tháng {intent.month}"
            sql += " GROUP BY kind"
            rows = get_db().execute(sql, params).fetchall()
            totals = {str(row["kind"]): int(row["total"] or 0) for row in rows}
            counts = {str(row["kind"]): int(row["n"] or 0) for row in rows}
            income_total = totals.get("income", 0)
            expense_total = totals.get("expense", 0)
            reply = f"{scope_label}: thu {_format_money(income_total)}, chi {_format_money(expense_total)}, còn lại {_format_money(income_total - expense_total)}."
            assistant = _save_chat(user_id, "assistant", reply, {"action": "query", "income": income_total, "expense": expense_total})
            return jsonify({"reply": reply, "assistant_message": assistant, "action": "query", "income": income_total, "expense": expense_total, "counts": counts, "used_ai": False})
        total, count = _query_total(user_id, intent)
        label = ALL_CATEGORIES.get(intent.category or "", "")
        subject = f" cho {label}" if label else ""
        kind_label = "thu" if intent.kind == "income" else "chi"
        scope_label = f"Tháng {intent.month}, " if intent.month else ""
        reply = f"{scope_label}tổng {kind_label}{subject} là {_format_money(total)} ({count} khoản)."
        assistant = _save_chat(user_id, "assistant", reply, {"action": "query", "total": total, "count": count})
        return jsonify({"reply": reply, "assistant_message": assistant, "action": "query", "total": total, "count": count, "used_ai": False})

    if intent.action == "export":
        query_parts: list[str] = []
        if intent.month:
            query_parts.append(f"month={intent.month}")
        if intent.category:
            query_parts.append(f"category={intent.category}")
        if intent.kind:
            query_parts.append(f"kind={intent.kind}")
        query = "&".join(query_parts)
        url = f"/api/finance/export.xlsx{('?' + query) if query else ''}"
        reply = f"Đã chuẩn bị Excel cho tháng {intent.month}." if intent.month else "Đã chuẩn bị Excel toàn bộ dữ liệu."
        assistant = _save_chat(user_id, "assistant", reply, {"action": "export", "download_url": url})
        return jsonify({"reply": reply, "assistant_message": assistant, "action": "export", "download_url": url, "used_ai": False})

    # Short follow-up after an auto-saved transaction: edit the latest transaction instead of creating a new one.
    if state.get("last_transaction_id") and _looks_like_last_edit(message) and intent.amount is None:
        fields = intent.to_dict()
        item = _update_last_transaction(user_id, str(state["last_transaction_id"]), fields)
        if item:
            reply = "Đã sửa khoản vừa rồi: " + _transaction_summary(item).removeprefix("Đã lưu: ")
            assistant = _save_chat(user_id, "assistant", reply, {"action": "updated", "transaction_id": item["id"]})
            return jsonify({"reply": reply, "assistant_message": assistant, "action": "updated", "transaction": item, "overview": _overview(user_id), "used_ai": False})

    draft = intent.to_dict()
    used_ai = False
    if intent.needs_clarification:
        _save_state(user_id, pending=draft)
        reply = intent.clarification or _pending_question(draft)
        assistant = _save_chat(user_id, "assistant", reply, {"action": "clarify", "draft": draft})
        return jsonify({"reply": reply, "assistant_message": assistant, "action": "clarify", "draft": draft, "used_ai": False})

    if intent.needs_ai:
        classification = _finance_ai_classifier(message, intent)
        if classification:
            used_ai = True
            for key in ("kind", "category", "vendor", "department", "project", "client", "payment_method", "confidence"):
                if classification.get(key) not in {None, ""}:
                    draft[key] = classification[key]
        if not classification or float(draft.get("confidence") or 0) < AI_AUTOSAVE_CONFIDENCE or not draft.get("category"):
            _save_state(user_id, pending=draft)
            reply = _pending_question(draft)
            assistant = _save_chat(user_id, "assistant", reply, {"action": "clarify", "draft": draft, "used_ai": used_ai})
            return jsonify({"reply": reply, "assistant_message": assistant, "action": "clarify", "draft": draft, "used_ai": used_ai})
        draft["review_status"] = "ai_auto"

    draft["confidence"] = float(draft.get("confidence") or 1.0)
    item = _insert_transaction(user_id, draft, source="finance_chat_ai" if used_ai else "finance_chat")
    _save_state(user_id, pending=None, last_transaction_id=item["id"])
    reply = _transaction_summary(item)
    assistant = _save_chat(user_id, "assistant", reply, {"action": "created", "transaction_id": item["id"], "used_ai": used_ai})
    return jsonify({"reply": reply, "assistant_message": assistant, "action": "created", "transaction": item, "overview": _overview(user_id), "used_ai": used_ai})


@bp.get("/api/finance/export.xlsx")
def finance_export_xlsx():
    user_id = _user_id()
    if not user_id:
        return _error("Bạn cần đăng nhập trước.", 401, "auth_required")
    raw_month = str(request.args.get("month") or "").strip()
    try:
        month = _parse_month(raw_month) if raw_month else ""
    except ValueError as exc:
        return _error(str(exc))
    category = str(request.args.get("category") or "").strip()
    kind = str(request.args.get("kind") or "").strip()

    sql = f"SELECT {_transaction_select()} FROM finance_transactions WHERE user_id = ?"
    params: list[Any] = [user_id]
    if month:
        start_date, end_date = _month_bounds(month)
        sql += " AND occurred_on BETWEEN ? AND ?"
        params.extend([start_date, end_date])
    if category:
        sql += " AND category = ?"
        params.append(category)
    if kind in {"income", "expense"}:
        sql += " AND kind = ?"
        params.append(kind)
    sql += " ORDER BY occurred_on ASC, created_at ASC"
    rows = [_serialize_transaction(row) for row in get_db().execute(sql, params).fetchall()]

    export_data = _build_export_data(rows)
    wb = Workbook()

    summary_ws = wb.active
    summary_ws.title = "Tổng quan"
    report_scope = f"THÁNG {month}" if month else "TOÀN BỘ DỮ LIỆU"
    summary_ws.append([f"BÁO CÁO THU CHI {report_scope}"])
    summary_ws["A1"].font = Font(bold=True, size=16)
    summary_ws.append([])
    summary_ws.append(["Tổng thu", export_data["summary"]["income"]])
    summary_ws.append(["Tổng chi", export_data["summary"]["expense"]])
    summary_ws.append(["Còn lại", export_data["summary"]["balance"]])
    for row_num in range(3, 6):
        summary_ws.cell(row=row_num, column=1).font = Font(bold=True)
        summary_ws.cell(row=row_num, column=2).number_format = '#,##0'
    summary_ws.append([])
    summary_ws.append(["CƠ CẤU CHI", "Số tiền", "Tỷ trọng"])
    for cell in summary_ws[7]:
        cell.font = Font(bold=True)
    total_expense = int(export_data["summary"]["expense"] or 0)
    for _key, label, amount in export_data["category_totals"]:
        share = (amount / total_expense) if total_expense else 0
        summary_ws.append([label, amount, share])
        summary_ws.cell(row=summary_ws.max_row, column=2).number_format = '#,##0'
        summary_ws.cell(row=summary_ws.max_row, column=3).number_format = '0.0%'
    summary_ws.column_dimensions["A"].width = 30
    summary_ws.column_dimensions["B"].width = 18
    summary_ws.column_dimensions["C"].width = 14

    ws = wb.create_sheet("Giao dịch")
    ws.append(export_data["detail_headers"])
    for cell in ws[1]:
        cell.font = Font(bold=True)
    for detail_row in export_data["detail_rows"]:
        ws.append(detail_row)
    ws.freeze_panes = "A2"
    for row_num in range(2, ws.max_row + 1):
        ws.cell(row=row_num, column=3).number_format = '#,##0'
    for column, width in {"A": 14, "B": 10, "C": 16, "D": 24, "E": 48, "F": 24, "G": 18, "H": 18, "I": 20, "J": 18}.items():
        ws.column_dimensions[column].width = width
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    filename = f"chi-tieu-{month}.xlsx" if month else "chi-tieu-toan-bo.xlsx"
    return send_file(
        buffer,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=filename,
    )


@bp.post("/api/finance/transaction")
def finance_add_transaction():
    user_id = _user_id()
    if not user_id:
        return _error("Bạn cần đăng nhập trước.", 401, "auth_required")
    payload = request.get_json(silent=True) or {}
    kind = str(payload.get("kind", "")).strip().lower()
    if kind not in {"income", "expense"}:
        return _error("Loại giao dịch không hợp lệ.")
    try:
        amount = _amount(payload.get("amount"))
    except ValueError as exc:
        return _error(str(exc))
    categories = INCOME_CATEGORIES if kind == "income" else EXPENSE_CATEGORIES
    category = str(payload.get("category", "")).strip()
    if category not in categories:
        return _error("Danh mục không hợp lệ.")
    note = " ".join(str(payload.get("note", "")).split()).strip()[:180]
    occurred_on = str(payload.get("occurred_on", "")).strip() or date.today().isoformat()
    try:
        datetime.strptime(occurred_on, "%Y-%m-%d")
    except ValueError:
        return _error("Ngày giao dịch không hợp lệ.")
    item = _insert_transaction(
        user_id,
        {"kind": kind, "amount": amount, "category": category, "note": note, "occurred_on": occurred_on, "confidence": 1.0, "review_status": "confirmed"},
        source="manual",
    )
    _save_state(user_id, last_transaction_id=item["id"])
    return jsonify({"transaction_id": item["id"], "overview": _overview(user_id)}), 201


@bp.delete("/api/finance/transaction/<transaction_id>")
def finance_delete_transaction(transaction_id: str):
    user_id = _user_id()
    if not user_id:
        return _error("Bạn cần đăng nhập trước.", 401, "auth_required")
    db = get_db()
    row = db.execute(
        "SELECT occurred_on FROM finance_transactions WHERE id = ? AND user_id = ?",
        (transaction_id, user_id),
    ).fetchone()
    if not row:
        return _error("Không tìm thấy giao dịch.", 404, "not_found")
    month = str(row["occurred_on"])[:7]
    db.execute("DELETE FROM finance_transactions WHERE id = ? AND user_id = ?", (transaction_id, user_id))
    db.commit()
    state = _assistant_state(user_id)
    if state.get("last_transaction_id") == transaction_id:
        _save_state(user_id, last_transaction_id="")
    return jsonify({"ok": True, "overview": _overview(user_id)})


@bp.post("/api/finance/plan")
def finance_save_plan():
    user_id = _user_id()
    if not user_id:
        return _error("Bạn cần đăng nhập trước.", 401, "auth_required")
    payload = request.get_json(silent=True) or {}
    try:
        month = _parse_month(payload.get("month"))
        monthly_income_target = _amount(payload.get("monthly_income_target", 0), allow_zero=True)
        budget_limit = _amount(payload.get("budget_limit", 0), allow_zero=True)
        saving_target = _amount(payload.get("saving_target", 0), allow_zero=True)
    except ValueError as exc:
        return _error(str(exc))
    now = _now()
    db = get_db()
    db.execute(
        """
        INSERT INTO finance_monthly_plans(user_id, month, monthly_income_target, budget_limit, saving_target, updated_at)
        VALUES (?, ?, ?, ?, ?, ?)
        ON CONFLICT(user_id, month) DO UPDATE SET
            monthly_income_target = excluded.monthly_income_target,
            budget_limit = excluded.budget_limit,
            saving_target = excluded.saving_target,
            updated_at = excluded.updated_at
        """,
        (user_id, month, monthly_income_target, budget_limit, saving_target, now),
    )
    db.commit()
    return jsonify({"ok": True, "overview": _overview(user_id, plan_month=month)})
